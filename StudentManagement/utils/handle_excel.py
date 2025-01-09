import openpyxl

# # Excelを読み込む
class ReadExcel():
    def __init__(self, file_path):
        # Excelをインポート
        self.workbook=openpyxl.load_workbook(file_path)
        # 最初のシートを取得
        self.worksheet=self.workbook.active

    def get_data(self):
        data=[]
        for row in self.worksheet.iter_rows():
            row_value=[]
            for cell in row:
                row_value.append(cell.value)
            data.append(row_value)
        return data
    
class WriteExcel():
    # この方法はローカルに書き込むものですが、ネットワークを通じてユーザーに送信し、
    # ユーザーがローカルにダウンロードできるようにする必要があります
    def __init__(self,file_path,data):
        self.file_path = file_path
        self.data = data
        # 新しい空のExcelワークブックを作成
        self.workbook=openpyxl.Workbook()
        self.worksheet=self.workbook.active

    def write_to_excel(self):
        # データをワークシートに追加
        for row in self.data:
            self.worksheet.append(row)
        self.workbook.save(self.file_path)

if __name__ == "__main__":
    import datetime
    # read_excel_obj=ReadExcel(r'D:\HAL\StudentManagement\1年1組学生情報.xlsx')
    # result=read_excel_obj.get_data()
    # print(result)
    data=[
    ['クラス', '名前', '学籍番号', '性別', '生年月日', '電話番号', '住所'],
    ['1年1組', '田中亮', 'G486190201508054369', '男', datetime.datetime(2015, 8, 5, 0, 0), '080-3644-9111', '東京都新宿区四谷町1丁目1-1'],
    ['1年1組', '鈴木舞', 'G533969201208026540', '女', datetime.datetime(2012, 8, 2, 0, 0), '090-4574-4893', '神奈川県横浜市中区桜通り2-3-4']
]
    write_excel_obj=WriteExcel(r'D:\HAL\StudentManagement\テスト.xlsx',data)
    write_excel_obj.write_to_excel()
